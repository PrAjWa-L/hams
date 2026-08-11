#!/usr/bin/env python3
"""Validate and import the multi-sheet HAMS asset workbook.

The default mode is a database-free dry run. Replacing the current asset
registry is deliberately explicit and transactional: validation is completed
before a database connection is opened, then deletion and insertion commit as
one unit.

Usage from ``backend/``::

    python scripts/import_assets_xlsx.py --file "../Asset Details .xlsx" --dry-run
    python scripts/import_assets_xlsx.py --file "../Asset Details .xlsx" \
        --replace-assets --confirm-delete-all-assets
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from sqlalchemy import delete, func, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.config import settings
from app.models.asset import Asset
from app.models.asset_assignment import AssetAssignment
from app.models.asset_category import AssetCategory
from app.models.maintenance_record import MaintenanceRecord


PLACEHOLDERS = {"", "-", "na", "n/a", "nil", "none", "nan", "no asset", "personal asset"}
SKIPPED_SHEETS = {"As per User list", "other"}
WORKSTATION_SHEETS = {
    "Cutis": ("Cutis", "Desktop"),
    "Chandra layout": ("Chandra Layout", "Laptop"),
    "HSR": ("HSR", "Laptop"),
    "Kochi": ("Kochi", "Laptop"),
}
INFRASTRUCTURE_SHEETS = {
    "NVR": "NVR",
    "NAS": "NAS",
    "Network": "Network",
    "AP": "AP",
    "Firewall": "Firewall",
}


@dataclass(slots=True)
class ImportIssue:
    severity: str
    sheet: str
    row: int
    message: str


@dataclass(slots=True)
class AssetRecord:
    sheet: str
    row: int
    name: str
    category: str
    branch: str
    brand: str | None = None
    model: str | None = None
    serial_number: str | None = None
    hostname: str | None = None
    ram: str | None = None
    hdd: str | None = None
    processor: str | None = None
    generation: str | None = None
    mac_address: str | None = None
    ip_address: str | None = None
    os_name: str | None = None
    os_activated: bool | None = None
    ms_office: str | None = None
    ms_office_activated: bool | None = None
    antivirus: str | None = None
    admin_login: bool | None = None
    floor: str | None = None
    label: str | None = None
    vendor_name: str | None = None
    notes: list[str] = field(default_factory=list)
    source_key: str = ""

    def __post_init__(self) -> None:
        if not self.source_key:
            digest = hashlib.sha256(
                f"{self.sheet}|{self.row}|{self.category}|{self.name}".encode("utf-8")
            ).hexdigest()[:20]
            self.source_key = f"XLSX-{digest.upper()}"


@dataclass(slots=True)
class ImportPlan:
    workbook: Path
    sheets: list[str]
    records: list[AssetRecord]
    issues: list[ImportIssue]

    @property
    def errors(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "error"]


def clean(value: Any) -> str | None:
    if value is None:
        return None
    value = re.sub(r"\s+", " ", str(value)).strip()
    return None if value.casefold() in PLACEHOLDERS else value


def clean_identity(value: Any) -> str | None:
    value = clean(value)
    if not value or value.casefold() in {"no asset", "personal asset"}:
        return None
    return value


def parse_bool(value: Any) -> bool | None:
    value = clean(value)
    if value is None:
        return None
    normalized = value.casefold().replace(" ", "").replace("-", "")
    if normalized in {"yes", "active", "activated", "true", "1", "available"}:
        return True
    if normalized in {"no", "notactive", "notactivated", "false", "0", "error"}:
        return False
    return None


def is_number(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    return bool(re.fullmatch(r"\d+(?:\.0+)?", str(value).strip())) if value is not None else False


def note(label: str, value: Any) -> str | None:
    value = clean(value)
    return f"{label}: {value}" if value else None


def infer_branch(value: Any, default: str = "Cutis") -> str:
    value = (clean(value) or "").casefold()
    if "hsr" in value:
        return "HSR"
    if "kochi" in value or "cochin" in value:
        return "Kochi"
    if "chandra" in value:
        return "Chandra Layout"
    if "cutis" in value:
        return "Cutis"
    return default


def workstation_records(ws, branch: str, default_category: str) -> tuple[list[AssetRecord], list[ImportIssue]]:
    records: list[AssetRecord] = []
    issues: list[ImportIssue] = []
    is_hsr = ws.title == "HSR"

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not is_number(row[0] if row else None):
            continue

        if is_hsr:
            user_name = clean(row[1])
            department = clean(row[2])
            hostname = clean_identity(row[3])
            processor, generation, ram, hdd = map(clean, row[4:8])
            ip_address, mac_address = clean(row[8]), clean(row[9])
            ms_office, ms_active = clean(row[10]), parse_bool(row[11])
            antivirus = "XDR/EPP" if parse_bool(row[12]) else None
            certificate = clean(row[14])
            admin_changed = parse_bool(row[15])
            os_name, os_active = clean(row[16]), parse_bool(row[17])
            floor = None
            category = "Server" if (department or "").casefold() == "server" else default_category
            physical_label = None
            remarks = None
        else:
            user_name = clean(row[1] if len(row) > 1 else None)
            department = clean(row[2] if len(row) > 2 else None)
            floor = clean(row[3] if len(row) > 3 else None)
            hostname = clean_identity(row[4] if len(row) > 4 else None)
            ram = clean(row[5] if len(row) > 5 else None)
            hdd = clean(row[6] if len(row) > 6 else None)
            processor = clean(row[7] if len(row) > 7 else None)
            generation = clean(row[8] if len(row) > 8 else None)
            mac_address = clean(row[9] if len(row) > 9 else None)
            ip_address = clean(row[10] if len(row) > 10 else None)
            os_name = clean(row[11] if len(row) > 11 else None)
            os_active = parse_bool(row[12] if len(row) > 12 else None)
            ms_office = clean(row[13] if len(row) > 13 else None)
            ms_active = parse_bool(row[14] if len(row) > 14 else None)
            admin_changed = parse_bool(row[15] if len(row) > 15 else None)
            xdr_index = 18 if len(row) > 18 else None
            antivirus = "XDR" if xdr_index is not None and parse_bool(row[xdr_index]) else None
            physical_label = clean(row[19] if len(row) > 19 else None)
            certificate = clean(row[20] if len(row) > 20 else None)
            raw_category = clean(row[21] if len(row) > 21 else None)
            category = raw_category.title() if raw_category else default_category
            remarks = clean(row[22] if len(row) > 22 else None)

        name = user_name or hostname
        if not name or not hostname:
            issues.append(ImportIssue("warning", ws.title, row_number, "Skipped placeholder or row without an asset hostname."))
            continue

        extra_notes = [
            note("Workbook department", department),
            note("Physical label", physical_label),
            note("Certificate uploaded", certificate),
            note("Remarks", remarks),
        ]
        records.append(AssetRecord(
            sheet=ws.title,
            row=row_number,
            name=name,
            category=category,
            branch=branch,
            hostname=hostname,
            ram=ram,
            hdd=hdd,
            processor=processor,
            generation=generation,
            mac_address=mac_address,
            ip_address=ip_address,
            os_name=os_name,
            os_activated=os_active,
            ms_office=ms_office,
            ms_office_activated=ms_active,
            antivirus=antivirus,
            admin_login=admin_changed,
            floor=floor,
            label=department,
            notes=[item for item in extra_notes if item],
        ))
    return records, issues


def printer_records(ws) -> tuple[list[AssetRecord], list[ImportIssue]]:
    records: list[AssetRecord] = []
    branch = "Cutis"
    coverage = "AMC"

    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = clean(row[0] if row else None)
        if first:
            normalized = first.casefold()
            if normalized in {"hsr", "kochi", "chandra layout"}:
                branch = infer_branch(first)
                continue
            if "not cover under amc" in normalized:
                coverage = "Not covered under AMC"
                continue
            if normalized == "amc":
                coverage = "AMC"
                continue
        if not is_number(row[0] if row else None):
            continue

        model = clean(row[1] if len(row) > 1 else None)
        serial = clean(row[2] if len(row) > 2 else None)
        department = clean(row[3] if len(row) > 3 else None)
        remarks = clean(row[4] if len(row) > 4 else None)
        if not model or not serial:
            continue
        records.append(AssetRecord(
            sheet=ws.title,
            row=row_number,
            name=f"{model} ({serial})",
            category="Printer",
            branch=branch,
            model=model,
            serial_number=serial,
            label=department,
            notes=[coverage, *([f"Location detail: {remarks}"] if remarks else [])],
        ))
    return records, []


def server_records(ws, category: str, branch: str = "Cutis") -> tuple[list[AssetRecord], list[ImportIssue]]:
    records: list[AssetRecord] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not is_number(row[0] if row else None):
            continue
        purpose = clean(row[4] if len(row) > 4 else None)
        hostname = clean_identity(row[5] if len(row) > 5 else None)
        name = purpose or hostname or clean(row[1] if len(row) > 1 else None)
        if not name:
            continue
        records.append(AssetRecord(
            sheet=ws.title,
            row=row_number,
            name=name,
            category=category,
            branch=branch,
            hostname=hostname,
            ram=clean(row[6] if len(row) > 6 else None),
            hdd=clean(row[7] if len(row) > 7 else None),
            processor=clean(row[8] if len(row) > 8 else None),
            generation=clean(row[9] if len(row) > 9 else None),
            mac_address=clean(row[11] if len(row) > 11 else None),
            ip_address=clean(row[12] if len(row) > 12 else None),
            os_name=clean(row[13] if len(row) > 13 else None),
            os_activated=parse_bool(row[14] if len(row) > 14 else None),
            ms_office=clean(row[15] if len(row) > 15 else None),
            ms_office_activated=parse_bool(row[16] if len(row) > 16 else None),
            antivirus="XDR/EPP" if parse_bool(row[19] if len(row) > 19 else None) else None,
            admin_login=parse_bool(row[17] if len(row) > 17 else None),
            floor=clean(row[3] if len(row) > 3 else None),
            label=clean(row[2] if len(row) > 2 else None),
            notes=[item for item in (
                note("Purpose", purpose),
                note("Cores", row[10] if len(row) > 10 else None),
                note("Remarks", row[22] if len(row) > 22 else None),
            ) if item],
        ))
    return records, []


def infrastructure_records(ws, category: str) -> tuple[list[AssetRecord], list[ImportIssue]]:
    records: list[AssetRecord] = []
    seen_names: Counter[str] = Counter()
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not is_number(row[0] if row else None):
            continue
        company = clean(row[1] if len(row) > 1 else None)
        department = clean(row[2] if len(row) > 2 else None)
        if not company:
            continue
        quantity_match = re.search(r"(\d+)\s*qty", department or "", re.IGNORECASE)
        quantity = int(quantity_match.group(1)) if quantity_match else 1
        branch = infer_branch(department)
        for unit in range(1, quantity + 1):
            base_name = company
            seen_names[f"{branch}|{base_name}"] += 1
            occurrence = seen_names[f"{branch}|{base_name}"]
            suffix = f" {unit}" if quantity > 1 else (f" {occurrence}" if occurrence > 1 else "")
            records.append(AssetRecord(
                sheet=ws.title,
                row=row_number,
                name=f"{base_name}{suffix}",
                category=category,
                branch=branch,
                brand=company,
                ip_address=clean(row[4] if len(row) > 4 else None),
                floor=clean(row[3] if len(row) > 3 else None),
                label=department,
                notes=[item for item in (
                    note("Physical label", row[5] if len(row) > 5 else None),
                    note("CCTV channels", row[6] if category == "NVR" and len(row) > 6 else None),
                    note("Remarks", row[7] if category == "NVR" and len(row) > 7 else (row[6] if len(row) > 6 else None)),
                ) if item],
                source_key="" if quantity == 1 else f"XLSX-{hashlib.sha256(f'{ws.title}|{row_number}|{unit}'.encode()).hexdigest()[:20].upper()}",
            ))
    return records, []


def rental_printer_records(ws) -> tuple[list[AssetRecord], list[ImportIssue]]:
    records: list[AssetRecord] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not is_number(row[0] if row else None):
            continue
        vendor, model, serial = clean(row[1]), clean(row[2]), clean(row[3])
        department = clean(row[4])
        if not model or not serial:
            continue
        records.append(AssetRecord(
            sheet=ws.title,
            row=row_number,
            name=f"{model} ({serial})",
            category="Rental Printer",
            branch=infer_branch(department),
            model=model,
            serial_number=serial,
            vendor_name=vendor,
            label=department,
            notes=[item for item in (note("DC number", row[5]), note("Workbook date", row[6])) if item],
        ))
    return records, []


def parse_workbook(path: Path) -> ImportPlan:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    records: list[AssetRecord] = []
    issues: list[ImportIssue] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        logical_name = sheet_name.strip()
        if logical_name in WORKSTATION_SHEETS:
            branch, default_category = WORKSTATION_SHEETS[logical_name]
            parsed, parsed_issues = workstation_records(ws, branch, default_category)
        elif logical_name == "Printer Details":
            parsed, parsed_issues = printer_records(ws)
        elif logical_name == "Server":
            parsed, parsed_issues = server_records(ws, "Server")
        elif logical_name == "Rental Server":
            parsed, parsed_issues = server_records(ws, "Rental Server")
        elif logical_name == "Rental Printer":
            parsed, parsed_issues = rental_printer_records(ws)
        elif logical_name in INFRASTRUCTURE_SHEETS:
            parsed, parsed_issues = infrastructure_records(ws, INFRASTRUCTURE_SHEETS[logical_name])
        elif logical_name in SKIPPED_SHEETS:
            continue
        else:
            issues.append(ImportIssue("warning", sheet_name, 0, "Unsupported sheet skipped."))
            continue
        records.extend(parsed)
        issues.extend(parsed_issues)

    records, duplicate_issues = remove_duplicates(records)
    issues.extend(duplicate_issues)
    if not records:
        issues.append(ImportIssue("error", "workbook", 0, "No importable asset rows were found."))
    workbook.close()
    return ImportPlan(path, list(workbook.sheetnames), records, issues)


def remove_duplicates(records: Iterable[AssetRecord]) -> tuple[list[AssetRecord], list[ImportIssue]]:
    accepted: list[AssetRecord] = []
    seen: dict[tuple[str, str], AssetRecord] = {}
    issues: list[ImportIssue] = []
    for record in records:
        keys = [("source", record.source_key.casefold())]
        for label, value in (
            ("serial number", record.serial_number),
            ("hostname", record.hostname),
            ("MAC address", record.mac_address),
        ):
            if value:
                keys.append((label, value.casefold()))
        duplicate = next(((label, seen[(label, value)]) for label, value in keys if (label, value) in seen), None)
        if duplicate:
            label, original = duplicate
            issues.append(ImportIssue(
                "warning",
                record.sheet,
                record.row,
                f"Duplicate {label}; skipped in favor of {original.sheet} row {original.row}.",
            ))
            continue
        accepted.append(record)
        for label, value in keys:
            seen[(label, value)] = record
    return accepted, issues


async def get_or_create_category(db: AsyncSession, name: str) -> AssetCategory:
    result = await db.execute(select(AssetCategory).where(
        func.lower(AssetCategory.name) == name.casefold(),
        AssetCategory.domain == "IT",
    ))
    category = result.scalar_one_or_none()
    if category is None:
        category = AssetCategory(id=uuid4(), name=name, domain="IT", is_active=True)
        db.add(category)
        await db.flush()
    return category


async def replace_assets(plan: ImportPlan) -> tuple[int, int, int, int]:
    engine = create_async_engine(settings.DATABASE_URL, future=True)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        async with factory() as db:
            old_assets = (await db.execute(select(func.count()).select_from(Asset))).scalar_one()
            old_assignments = (await db.execute(select(func.count()).select_from(AssetAssignment))).scalar_one()
            old_maintenance = (await db.execute(select(func.count()).select_from(MaintenanceRecord))).scalar_one()

            # Explicit dependent deletes make behavior safe even if an older local
            # database was created without the current ON DELETE CASCADE clauses.
            await db.execute(delete(AssetAssignment))
            await db.execute(delete(MaintenanceRecord))
            await db.execute(update(Asset).values(parent_asset_id=None))
            await db.execute(delete(Asset))
            await db.execute(text("CREATE SEQUENCE IF NOT EXISTS asset_seq_it START 1 INCREMENT 1"))
            await db.execute(text("ALTER SEQUENCE asset_seq_it RESTART WITH 1"))

            categories: dict[str, AssetCategory] = {}
            for name in sorted({record.category for record in plan.records}):
                categories[name] = await get_or_create_category(db, name)

            for sequence, record in enumerate(plan.records, start=1):
                source_note = f"Import source: {plan.workbook.name} · {record.sheet} row {record.row}"
                asset = Asset(
                    id=uuid4(),
                    asset_id=f"HAMS-IT-{sequence:05d}",
                    name=record.name[:200],
                    category_id=categories[record.category].id,
                    brand=record.brand,
                    model=record.model,
                    serial_number=record.serial_number,
                    barcode=record.source_key,
                    hostname=record.hostname,
                    ram=record.ram,
                    hdd=record.hdd,
                    processor=record.processor,
                    generation=record.generation,
                    mac_address=record.mac_address,
                    ip_address=record.ip_address,
                    os_name=record.os_name,
                    os_activated=record.os_activated,
                    ms_office=record.ms_office,
                    ms_office_activated=record.ms_office_activated,
                    antivirus=record.antivirus,
                    admin_login=record.admin_login,
                    floor=record.floor,
                    location_notes=record.branch,
                    label=record.label,
                    vendor_name=record.vendor_name,
                    notes="\n".join([source_note, *record.notes]),
                    status="available",
                    is_shared=False,
                    is_deleted=False,
                )
                db.add(asset)
            await db.flush()
            await db.execute(
                text("SELECT setval('asset_seq_it', :last_value, true)"),
                {"last_value": len(plan.records)},
            )
            await db.commit()
            return old_assets, old_assignments, old_maintenance, len(plan.records)
    except Exception:
        # The session context rolls back the deletion and insertions together.
        raise
    finally:
        await engine.dispose()


def print_plan(plan: ImportPlan) -> None:
    print(f"Workbook: {plan.workbook}")
    print(f"Sheets detected ({len(plan.sheets)}): {', '.join(plan.sheets)}")
    print(f"Importable assets: {len(plan.records)}")
    print("By branch:")
    for name, count in sorted(Counter(record.branch for record in plan.records).items()):
        print(f"  {name}: {count}")
    print("By type:")
    for name, count in sorted(Counter(record.category for record in plan.records).items()):
        print(f"  {name}: {count}")
    if plan.issues:
        print(f"Issues ({len(plan.issues)}):")
        for issue in plan.issues:
            location = f"{issue.sheet} row {issue.row}" if issue.row else issue.sheet
            print(f"  {issue.severity.upper():7} {location}: {issue.message}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, type=Path, help="Path to the XLSX workbook")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Validate only (default)")
    mode.add_argument("--replace-assets", action="store_true", help="Delete all assets and import this workbook")
    parser.add_argument(
        "--confirm-delete-all-assets",
        action="store_true",
        help="Required with --replace-assets",
    )
    return parser


async def async_main(args: argparse.Namespace) -> int:
    workbook_path = args.file.expanduser().resolve()
    if not workbook_path.is_file():
        print(f"ERROR: Workbook not found: {workbook_path}", file=sys.stderr)
        return 2
    if workbook_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        print("ERROR: Only .xlsx and .xlsm workbooks are supported.", file=sys.stderr)
        return 2

    plan = parse_workbook(workbook_path)
    print_plan(plan)
    if plan.errors:
        print("ERROR: Validation failed; the database was not changed.", file=sys.stderr)
        return 1
    if not args.replace_assets:
        print("Dry run complete. The database was not changed.")
        return 0
    if not args.confirm_delete_all_assets:
        print("ERROR: --confirm-delete-all-assets is required with --replace-assets.", file=sys.stderr)
        return 2

    old_assets, old_assignments, old_maintenance, created = await replace_assets(plan)
    print(
        "Replacement committed: "
        f"deleted {old_assets} assets, {old_assignments} assignments, "
        f"{old_maintenance} maintenance records; imported {created} assets."
    )
    return 0


def main() -> int:
    return asyncio.run(async_main(build_parser().parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
