from __future__ import annotations

from pathlib import Path

import openpyxl

from scripts.import_assets_xlsx import parse_bool, parse_workbook


def build_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    cutis = workbook.active
    cutis.title = "Cutis  "
    cutis.append([
        "Sl.NO", "User Name", "Department", "Floor", "Hostname", "RAM", "HDD",
        "Processor", "Generation", "MAC Address", "IP Address", "Windows", "Activated",
        "MS office", "Activated", "Administrator name changed", "Hostname changed",
        "Removal other antivrues", "XDR Installed", "Label", "Certificate",
        "Type of Asset", "Remarks",
    ])
    cutis.append([
        1, "Fixture Desktop", "IT", "2nd", "fixture-host", "8 GB", "512 GB", "i5",
        "12th", "AA:BB:CC:DD:EE:FF", "192.0.2.10", "Windows 11", "Active", "Office",
        "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Desktop", None,
    ])

    chandra = workbook.create_sheet("Chandra layout")
    chandra.append(list(cutis.iter_rows(min_row=1, max_row=1, values_only=True))[0])
    chandra.append([
        1, "Duplicate", "IT", "1st", "fixture-host", "8 GB", "512 GB", "i5", "12th",
        "11:22:33:44:55:66", "192.0.2.11", "Windows 11", "Active", "Office", "Yes",
        "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Laptop", None,
    ])

    printers = workbook.create_sheet("Printer Details ")
    printers.append([None, None, None, None, None])
    printers.append(["AMC", None, None, None, None])
    printers.append(["SI NO", "Model", "Serial Number", "Department", None])
    printers.append([1, "Fixture Printer", "SERIAL-001", "IT", None])
    printers.append(["HSR", None, None, None, None])
    printers.append([1, "Fixture HSR Printer", "SERIAL-002", "HSR", "Reception"])

    workbook.create_sheet("As per User list").append(["Username"])
    workbook.save(path)


def test_parser_supports_trailing_sheet_spaces_and_deduplicates(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fixture.xlsx"
    build_fixture(workbook_path)

    plan = parse_workbook(workbook_path)

    assert len(plan.records) == 3
    assert {(record.category, record.branch) for record in plan.records} == {
        ("Desktop", "Cutis"),
        ("Printer", "Cutis"),
        ("Printer", "HSR"),
    }
    assert any("Duplicate hostname" in issue.message for issue in plan.issues)
    assert not plan.errors


def test_boolean_normalization() -> None:
    assert parse_bool(" Active ") is True
    assert parse_bool("not- active") is False
    assert parse_bool("Error") is False
    assert parse_bool(None) is None
