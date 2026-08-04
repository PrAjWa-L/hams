from __future__ import annotations

import io
from datetime import datetime
from typing import List
from uuid import UUID

from fastapi import APIRouter, Depends, Query, Response, status, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.dependencies.auth import AuthUser
from app.api.v1.dependencies.pagination import PagedResponse, Pagination
from app.core.rbac import Permission, require_permissions
from app.core.storage import get_presigned_url
from app.db.session import get_db
from app.schemas.asset import (
    AssetCreateRequest,
    AssetListItem,
    AssetResponse,
    AssetRetireRequest,
    AssetTransferRequest,
    AssetUpdateRequest,
    DocumentResponse,
    ImportSummary,
)
from app.schemas.response import APIResponse
from app.services.asset_service import AssetService
from app.services.document_service import DocumentService

router = APIRouter(tags=["Assets"])


@router.post(
    "",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_201_CREATED,
    summary="Register a new asset",
    dependencies=[Depends(require_permissions(Permission.ASSET_CREATE))],
)
async def create_asset(
    payload: AssetCreateRequest,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.create(
        payload, actor_id=current_user.id, actor_role=current_user.role
    )
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.get(
    "",
    response_model=PagedResponse[AssetListItem],
    status_code=status.HTTP_200_OK,
    summary="List assets with filters — domain-scoped by role",
)
async def list_assets(
    pagination: Pagination,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
    category_id: UUID | None = Query(None),
    category_name: str | None = Query(None),
    department_id: UUID | None = Query(None),
    status: str | None = Query(None),
    domain: str | None = Query(None, description="IT or FACILITY — COO only"),
    floor: str | None = Query(None),
    site: str | None = Query(None, description="Branch filter: Cutis, HSR, Kochi"),
    search: str | None = Query(None, max_length=100),
    warranty_expiring_days: int | None = Query(
        None, ge=1, le=365, description="Assets with warranty expiring within N days"
    ),
):
    svc = AssetService(db)
    assets, total = await svc.list_assets(
        actor_role=current_user.role,
        offset=pagination.offset,
        limit=pagination.limit,
        category_id=category_id,
        category_name=category_name,
        department_id=department_id,
        status=status,
        domain=domain,
        floor=floor,
        site=site,
        search=search,
        warranty_expiring_days=warranty_expiring_days,
    )
    return PagedResponse.build(
        items=[AssetListItem.model_validate(a) for a in assets],
        total=total,
        params=pagination,
    )


@router.get(
    "/export",
    summary="Export all assets to Excel (.xlsx)",
    dependencies=[Depends(require_permissions(Permission.ASSET_READ))],
)
async def export_assets(
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
    status: str | None = Query(None),
    domain: str | None = Query(None),
    category_id: UUID | None = Query(None),
    category_name: str | None = Query(None),
    site: str | None = Query(None),
    search: str | None = Query(None),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    svc = AssetService(db)
    assets, _ = await svc.list_assets(
        actor_role=current_user.role,
        offset=0,
        limit=10000,   # fetch all
        status=status,
        domain=domain,
        category_id=category_id,
        category_name=category_name,
        site=site,
        search=search,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    HEADERS = [
        "Asset ID", "Name", "Category", "Domain", "Status",
        "Brand", "Model", "Serial Number", "Hostname",
        "RAM", "HDD", "Processor", "OS",
        "Floor", "Department", "Label",
        "Purchase Date", "Purchase Cost", "PO Reference",
        "Vendor", "Purchased From",
        "Warranty End", "AMC End",
        "Notes",
    ]

    # Header row styling
    header_fill = PatternFill("solid", fgColor="0F4C5C")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Status colour map
    STATUS_COLORS = {
        "available":        "E8FCD4",
        "assigned":         "DCE9FF",
        "under_maintenance":"FFF5CC",
        "retired":          "FDE7E7",
    }

    for row_idx, asset in enumerate(assets, 2):
        row_data = [
            asset.asset_id,
            asset.name,
            asset.category.name if asset.category else "",
            getattr(asset, "domain", ""),
            asset.status,
            getattr(asset, "brand", "") or "",
            getattr(asset, "model", "") or "",
            getattr(asset, "serial_number", "") or "",
            getattr(asset, "hostname", "") or "",
            getattr(asset, "ram", "") or "",
            getattr(asset, "hdd", "") or "",
            getattr(asset, "processor", "") or "",
            getattr(asset, "os_name", "") or "",
            getattr(asset, "floor", "") or "",
            asset.department.name if getattr(asset, "department", None) else "",
            getattr(asset, "label", "") or "",
            str(asset.purchase_date) if getattr(asset, "purchase_date", None) else "",
            str(asset.purchase_cost) if getattr(asset, "purchase_cost", None) else "",
            getattr(asset, "po_reference", "") or "",
            getattr(asset, "vendor_name", "") or "",
            getattr(asset, "purchased_from", "") or "",
            str(asset.warranty_end) if getattr(asset, "warranty_end", None) else "",
            str(asset.amc_end) if getattr(asset, "amc_end", None) else "",
            getattr(asset, "notes", "") or "",
        ]
        color = STATUS_COLORS.get(asset.status, "FFFFFF")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")

    # Column widths
    COL_WIDTHS = [12,28,16,8,14,14,16,18,20,8,10,22,16,10,18,12,14,12,14,18,18,14,12,35]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"HAMS_Assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/by-code/{asset_code}",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_200_OK,
    summary="Get asset by asset code (e.g. HAMS-IT-00001) — used by QR scanner",
)
async def get_asset_by_code(
    asset_code: str,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.get_by_asset_code(asset_code, actor_role=current_user.role)
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.get(
    "/{asset_id}",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_200_OK,
    summary="Get asset by UUID",
)
async def get_asset(
    asset_id: UUID,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.get_by_id(asset_id, actor_role=current_user.role)
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.patch(
    "/{asset_id}",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_200_OK,
    summary="Update asset details",
    dependencies=[Depends(require_permissions(Permission.ASSET_UPDATE))],
)
async def update_asset(
    asset_id: UUID,
    payload: AssetUpdateRequest,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.update(
        asset_id, payload, actor_id=current_user.id, actor_role=current_user.role
    )
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.post(
    "/{asset_id}/retire",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_200_OK,
    summary="Retire an asset",
    dependencies=[Depends(require_permissions(Permission.ASSET_RETIRE))],
)
async def retire_asset(
    asset_id: UUID,
    payload: AssetRetireRequest,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.retire(
        asset_id, payload, actor_id=current_user.id, actor_role=current_user.role
    )
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.post(
    "/{asset_id}/transfer",
    response_model=APIResponse[AssetResponse],
    status_code=status.HTTP_200_OK,
    summary="Transfer asset to another department",
    dependencies=[Depends(require_permissions(Permission.ASSET_TRANSFER))],
)
async def transfer_asset(
    asset_id: UUID,
    payload: AssetTransferRequest,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.transfer(
        asset_id, payload, actor_id=current_user.id, actor_role=current_user.role
    )
    return APIResponse.ok(data=AssetResponse.model_validate(asset))


@router.get(
    "/{asset_id}/qr",
    status_code=status.HTTP_200_OK,
    summary="Get presigned QR code URL for an asset",
)
async def get_asset_qr(
    asset_id: UUID,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    svc = AssetService(db)
    asset = await svc.get_by_id(asset_id, actor_role=current_user.role)
    if not asset.qr_code_url:
        from app.core.exceptions import NotFoundException
        raise NotFoundException(detail="QR code not yet generated for this asset")
    url = get_presigned_url(asset.qr_code_url, expires_seconds=300)
    return APIResponse.ok(data={"qr_url": url, "expires_in": 300})


@router.get(
    "/{asset_id}/documents",
    response_model=APIResponse[List[DocumentResponse]],
    status_code=status.HTTP_200_OK,
    summary="List documents attached to an asset",
)
async def list_asset_documents(
    asset_id: UUID,
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
):
    # Verify asset access first
    asset_svc = AssetService(db)
    await asset_svc.get_by_id(asset_id, actor_role=current_user.role)

    doc_svc = DocumentService(db)
    docs = await doc_svc.list_for_entity("asset", asset_id)
    return APIResponse.ok(data=[DocumentResponse.model_validate(d) for d in docs])


# ── CSV Import ─────────────────────────────────────────────────────────────────

from io import StringIO
from typing import Any, Dict



@router.post(
    "/import",
    response_model=APIResponse[ImportSummary],
    status_code=status.HTTP_200_OK,
    summary="Bulk-import IT assets from a Seqrite SystemAndHardwareDetails CSV export",
    dependencies=[Depends(require_permissions(Permission.ASSET_CREATE))],
)
async def import_assets_csv(
    current_user: AuthUser,
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
):
    if not file.filename or not file.filename.lower().endswith(".csv"):
        from app.core.exceptions import BadRequestException
        raise BadRequestException(detail="Only .csv files are accepted")
    content = await file.read()
    svc = AssetService(db)
    summary = await svc.import_csv(
        content, actor_id=current_user.id, actor_role=current_user.role
    )
    return APIResponse.ok(data=summary)